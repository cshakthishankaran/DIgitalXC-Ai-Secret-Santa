
import openpyxl
import subprocess
import sys
import random
from openpyxl.styles import Alignment, PatternFill, Font
import os



class SecretSantaGenerator :

    pass

    def __init__(self, employee_list_filepath,previous_secret_santas_filepath):
       
      self.setup_virtual_environment()
      self.employees = self.read_from_xlsx(employee_list_filepath)
      self.previous_secret_santas = self.read_from_xlsx(previous_secret_santas_filepath)      
      


    def alot_new_secret_santas(self):
       
      """
        with the extracted data the new secret santas are generated.
        :param employees list & previous year's secret santas
        :return list of new secret santas after applying all the required constraints

      """
      print("Findinh new Secret Santas .....")
      employee_list = self.employees[:]
      new_secret_santas = []
      self.previous_secret_santas = {(entry['Employee_EmailID'], entry['Secret_Child_EmailID']) 
                           for entry in self.previous_secret_santas}
      for employee in self.employees:
          # Create a list of possible children (exclude the employee themselves and previous pairings)
          possible_children = [e for e in employee_list 
                                if e['Employee_EmailID'] != employee['Employee_EmailID'] and
                                (employee['Employee_EmailID'], e['Employee_EmailID']) not in self.previous_secret_santas]

          if not possible_children:
              raise NotImplementedError(f"No Secret Child available for {employee['Employee_Name']}")
          
          # Randomly choose a secret child from the possible candidates
          secret_child = random.choice(possible_children)
          employee_list.remove(secret_child)  # Ensure the secret child is not reused

          new_secret_santas.append({
              'Employee_Name': employee['Employee_Name'],
              'Employee_EmailID': employee['Employee_EmailID'],
              'Secret_Child_Name': secret_child['Employee_Name'],
              'Secret_Child_EmailID': secret_child['Employee_EmailID']
          })

      return new_secret_santas


      
    def setup_virtual_environment(self):
      """
      Checks if the virtual environment exists. If not, creates it 
      installs the required dependencies.

      """
      env_dir = 'env'
      
      # Checking if the virtual environment already exists
      if not os.path.exists(env_dir):
          print("No Virtual Environment available. Creating a virtual environment , sit back and relax ...")          
          
          # Create virtual environment with kernel command
          subprocess.check_call([sys.executable, '-m', 'venv', env_dir])
          print("Virtual environment created successfully.")

          

      
      requirements_file = 'requirements.txt'

       
      if os.path.exists(requirements_file):
          # Installing dependencies from requirements.txt
          print("Installing dependencies from requirements.txt...")
          subprocess.check_call([os.path.join(env_dir, 'Scripts', 'pip') if os.name == 'nt' else os.path.join(env_dir, 'bin', 'pip'), 'install', '-r', requirements_file])
          print("Dependencies installed successfully.")
      else:
          print(f"No {requirements_file} found. Please make sure you have the required dependencies file.")


    def read_from_xlsx(self, file_path):
              
        """
        Examines the xlsx file from the file path provided
        
        :param file_path: Path to the Excel file
        :return: List of dictionaries representing the data in the Excel file

        """
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]  # Read the header row
            data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Read from the second row
                data.append(dict(zip(headers, row)))
            return data
        except FileNotFoundError:
            print(f"Error: The file {file_path} was not found.")
            return []
        
    def write_to_xlsx(self, output_file, secret_santas):
      """
      Writes the new secret santa info to a new Excel (.xlsx) file.

      :param output_file: Path to the output Excel file (.xlsx).
      """
      try:
          # Create a new workbook and select the active sheet
          workbook = openpyxl.Workbook()
          sheet = workbook.active

          # Let's highlight the header
          header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow fill
          header_font = Font(bold=True)  # Bold font for headers
          center_alignment = Alignment(horizontal='center', vertical='center')
          # Write the header row
          sheet.append(['Employee_Name', 'Employee_EmailID', 'Secret_Child_Name', 'Secret_Child_EmailID'])

          # Write the secret_santa_pair rows
          for secret_santa_pair in secret_santas:
              sheet.append([
                  secret_santa_pair['Employee_Name'],
                  secret_santa_pair['Employee_EmailID'],
                  secret_santa_pair['Secret_Child_Name'],
                  secret_santa_pair['Secret_Child_EmailID']
              ])

          # Save the workbook to the specified output file
          workbook.save(output_file)
          print(f"Secret santas saved to {output_file}")
      
      except Exception as e:
          print(f"Error writing to file {output_file}: {str(e)}")
    


def letsPlaySecretSanta ():
  
  """
    Below are the files that need to be analysed for playing Secret Santa game.

    Files :

      1.  Employee_List.xlsx
      2.  Secret-Santa-Game-Result-2023.xlsx

  """

  employee_list_filepath = "Employee-List.xlsx"
  previous_secret_santas_filepath = "Secret-Santa-Game-Result-2023.xlsx"
  new_secret_santas_filepath = "Secret-Santa-Game-Result-2024.xlsx"

  secret_santa_generator = SecretSantaGenerator(employee_list_filepath,previous_secret_santas_filepath)  
  secret_santas = secret_santa_generator.alot_new_secret_santas()
  secret_santa_generator.write_to_xlsx("secret-santa-2024.xlsx",secret_santas)





if __name__ == "__main__" :
  letsPlaySecretSanta()
